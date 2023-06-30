import os
import sqlite3

import openpyxl  # Создаем файл Excel и записываем данные
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import StatesGroup, State
from openpyxl.utils import get_column_letter

from keyboards.admin_keyboards import admin_panel_keyboard, order_status
from system.dispatcher import dp, bot

# Подключение к базе данных SQLite
conn = sqlite3.connect('setting/orders.db')
cursor = conn.cursor()
# Создаем таблицу, если она не существует
cursor.execute('''CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER,
                                                    first_name TEXT, last_name TEXT, username TEXT, date TEXT)''')
# Создаем таблицу order_status
cursor.execute('''CREATE TABLE IF NOT EXISTS order_status (order_number, in_processing, sent, cancelled, refund)''')
conn.commit()


@dp.callback_query_handler(lambda c: c.data == 'admin')
async def admin_panel__handler(callback_query: types.CallbackQuery):
    """Калькулятор цены исходя из типа валюты"""
    message_text_clothing = "Админ панель для работы с ботом"
    admin_keyboard = admin_panel_keyboard()
    # Клавиатура для выбора товара
    await bot.send_message(callback_query.from_user.id, message_text_clothing, reply_markup=admin_keyboard,
                           parse_mode=types.ParseMode.HTML)


@dp.callback_query_handler(lambda c: c.data == 'check_bot_users')
async def admin_panel_handler(callback_query: types.CallbackQuery):
    # Проверяем, является ли пользователь, который вызывает команду, администратором
    if callback_query.from_user.id not in [5837917794, 1062323239]:  # Если это не админ, то выводим предупреждение
        await bot.send_message(callback_query.from_user.id, 'У вас нет доступа к этой команде.')
        return

    # Получаем данные всех пользователей из базы данных
    cursor.execute('SELECT * FROM users')
    data = cursor.fetchall()

    wb = openpyxl.Workbook()
    sheet = wb.active

    # Записываем заголовки
    headers = ['ID', 'User ID', 'First Name', 'Last Name', 'Username', 'Date']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Записываем данные пользователей
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_data

    # Сохраняем файл Excel
    wb.save('users.xlsx')

    # Отправляем файл пользователю
    with open('users.xlsx', 'rb') as file:
        await bot.send_document(callback_query.from_user.id, file, caption='Данные пользователей в формате Excel')

    # Удаляем файл Excel
    import os
    os.remove('users.xlsx')


def create_excel_file(orders):
    """Функция для создания файла Excel с данными заказов"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'User ID'
    sheet['B1'] = 'Имя'
    sheet['C1'] = 'Номер телефона'
    sheet['D1'] = 'Ссылка на товар'
    sheet['E1'] = 'Размер'
    sheet['F1'] = 'Цвет товара'
    sheet['G1'] = 'Цена на товар в юанях'
    sheet['H1'] = 'Номер заказа'
    sheet['I1'] = 'Дата заказа'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # User ID
        sheet.cell(row=index, column=2).value = order[1]  # Имя
        sheet.cell(row=index, column=3).value = order[2]  # Номер телефона
        sheet.cell(row=index, column=4).value = order[3]  # Ссылка
        sheet.cell(row=index, column=5).value = order[4]  # Размер
        sheet.cell(row=index, column=6).value = order[5]  # Цвет
        sheet.cell(row=index, column=7).value = order[6]  # Цена в юанях
        sheet.cell(row=index, column=8).value = order[7]  # Номер заказа
        sheet.cell(row=index, column=9).value = order[8]  # Дата заказа

    return workbook


@dp.callback_query_handler(lambda c: c.data == 'unload_orders')
async def admin_panel_handler(callback_query: types.CallbackQuery):
    """Обработчик команды для выгрузки данных в Excel"""
    if callback_query.from_user.id not in [5837917794, 1062323239]:  # Если это не админ, то выводим предупреждение
        await bot.send_message(callback_query.from_user.id, 'У вас нет доступа к этой команде.')
        return
    # Получение данных из базы данных
    cursor.execute("SELECT * FROM orders")
    orders = cursor.fetchall()
    # Создание файла Excel
    workbook = create_excel_file(orders)
    # Сохранение файла
    filename = 'orders.xlsx'
    workbook.save(filename)
    # Отправка файла пользователю
    with open(filename, 'rb') as file:
        await bot.send_document(callback_query.from_user.id, file, caption='Данные пользователей в формате Excel')
    # Удаление файла
    os.remove(filename)


class OrderStatus(StatesGroup):
    """Статусы заказа"""
    waiting_for_order_number = State()


@dp.callback_query_handler(lambda c: c.data == 'change_order_status')
async def change_order_status_handler(callback_query: types.CallbackQuery):
    # Проверяем, является ли пользователь, который вызывает команду, администратором
    if callback_query.from_user.id not in [5837917794, 1062323239]:
        await callback_query.answer('У вас нет доступа к этой команде.', show_alert=True)
        return

    message_text_clothing = "Выберите статус заказа:"
    order_status_keyboard = order_status()
    # Клавиатура для выбора статуса заказа
    await bot.send_message(callback_query.from_user.id, message_text_clothing, reply_markup=order_status_keyboard,
                           parse_mode=types.ParseMode.HTML)


@dp.callback_query_handler(lambda c: c.data in ['in_processing', 'sent', 'cancelled', 'refund'])
async def set_order_status_handler(callback_query: types.CallbackQuery):
    # Проверяем, является ли пользователь, который вызывает команду, администратором
    if callback_query.from_user.id not in [5837917794, 1062323239]:
        await callback_query.answer('У вас нет доступа к этой команде.', show_alert=True)
        return

    # Получаем статус заказа, выбранный администратором
    order_status = callback_query.data

    # Запрашиваем у админа номер заказа
    await bot.send_message(callback_query.from_user.id, 'Введите номер заказа:')
    await OrderStatus.waiting_for_order_number.set()

    # Сохраняем статус заказа, чтобы использовать его при обработке номера заказа
    await dp.current_state().update_data(order_status=order_status)


@dp.message_handler(state=OrderStatus.waiting_for_order_number)
async def process_order_number(message: types.Message, state: FSMContext):
    order_number = message.text

    # Получаем сохраненный статус заказа из состояния
    data = await state.get_data()
    order_status = data.get('order_status')

    # Обновляем соответствующий статус заказа в базе данных
    cursor.execute(f"UPDATE order_status SET {order_status} = ? WHERE order_number = ?", (True, order_number))

    # Обнуляем другие статусы заказа
    for column in ['in_processing', 'sent', 'cancelled', 'refund']:
        if column != order_status:
            cursor.execute(f"UPDATE order_status SET {column} = ? WHERE order_number = ?", (False, order_number))

    conn.commit()

    # Проверяем, сколько строк было изменено
    if cursor.rowcount > 0:
        # Отправляем подтверждение об успешном обновлении статуса
        await message.answer(f"Статус заказа с номером {order_number} успешно обновлен.")
    else:
        # Если ни одна строка не была изменена, возможно, указанный номер заказа не существует
        await message.answer(f"Заказ с номером {order_number} не найден.")

    await state.finish()


def admin_handler():
    """Регистрируем handlers для администратора"""
    dp.register_callback_query_handler(change_order_status_handler, lambda c: c.data == 'change_order_status')
    dp.register_callback_query_handler(set_order_status_handler,
                                       lambda c: c.data in ['in_processing', 'sent', 'cancelled', 'refund'])
    dp.register_message_handler(process_order_number, state=OrderStatus.waiting_for_order_number)
