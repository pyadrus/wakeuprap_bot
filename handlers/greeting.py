# Импортируем модуль для работы с базой данных SQLite
import os
import sqlite3
import datetime

import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import StatesGroup, State

from keyboards.greeting_keyboards import greeting_keyboards
from messages.greeting_post import greeting_post
from system.dispatcher import dp, bot

# Подключение к базе данных SQLite
conn = sqlite3.connect('setting/orders.db')
cursor = conn.cursor()
# Создание таблицы для заказов
# cursor.execute('''CREATE TABLE IF NOT EXISTS orders (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER,
#                                                      name TEXT, phone TEXT, link TEXT, size TEXT, color TEXT,
#                                                      price TEXT, date TEXT)''')
# Создаем таблицу, если она не существует
cursor.execute('''CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER,
                                                    first_name TEXT, last_name TEXT, username TEXT, date TEXT)''')
conn.commit()


@dp.message_handler(commands=['start'])
async def greeting(message: types.Message, state: FSMContext):
    """Обработчик команды /start, он же пост приветствия"""
    await state.reset_state()

    # Получаем текущую дату и время
    current_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Записываем данные пользователя в базу данных
    cursor.execute('''INSERT INTO users (user_id, first_name, last_name, username, date) VALUES (?, ?, ?, ?, ?)''',
                   (message.from_user.id, message.from_user.first_name, message.from_user.last_name,
                    message.from_user.username, current_date))
    conn.commit()

    print(f'Привет! нажали на кнопку /start {message.from_user.id, message.from_user.username, current_date}')

    keyboards_greeting = greeting_keyboards()
    # Клавиатура для Калькулятора цен или Контактов
    await message.reply(greeting_post, reply_markup=keyboards_greeting, disable_web_page_preview=True,
                        parse_mode=types.ParseMode.HTML)


# Обработчик команды /export
@dp.message_handler(commands=['export_user'])
async def export_command(message: types.Message):
    # Проверяем, является ли пользователь, который вызывает команду, администратором
    if message.from_user.id not in [5837917794, 5958542955]:
        await message.reply('У вас нет доступа к этой команде.')
        return
    # Получаем данные всех пользователей из базы данных
    cursor.execute('SELECT * FROM users')
    data = cursor.fetchall()
    # Создаем файл Excel и записываем данные
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    sheet = wb.active
    # Записываем заголовки
    headers = ['ID', 'User ID', 'First Name', 'Last Name', 'Username', 'Date']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header
    # Записываем данные пользователей
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_data
    # Сохраняем файл Excel
    wb.save('users.xlsx')
    # Отправляем файл пользователю
    with open('users.xlsx', 'rb') as file:
        await bot.send_document(message.from_user.id, file, caption='Данные пользователей в формате Excel')
    # Удаляем файл Excel
    import os
    os.remove('users.xlsx')


# Функция для создания файла Excel с данными заказов
def create_excel_file(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    # sheet['A1'] = 'ID'
    sheet['A1'] = 'User ID'
    sheet['B1'] = 'Имя'
    sheet['C1'] = 'Номер телефона'
    sheet['D1'] = 'Ссылка на товар'
    sheet['E1'] = 'Размер'
    sheet['F1'] = 'Цвет товара'
    sheet['G1'] = 'Цена на товар в юанях'
    sheet['H1'] = 'Номер заказа'
    sheet['I1'] = 'Дата заказа'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        # sheet.cell(row=index, column=1).value = order[0]  # ID
        sheet.cell(row=index, column=1).value = order[0]  # User ID
        sheet.cell(row=index, column=2).value = order[1]  # Имя
        sheet.cell(row=index, column=3).value = order[2]  # Номер телефона
        sheet.cell(row=index, column=4).value = order[3]  # Ссылка
        sheet.cell(row=index, column=5).value = order[4]  # Размер
        sheet.cell(row=index, column=6).value = order[5]  # Цвет
        sheet.cell(row=index, column=7).value = order[6]  # Цена в юанях
        sheet.cell(row=index, column=8).value = order[7]  # Номер заказа
        sheet.cell(row=index, column=9).value = order[8]  # Дата заказа

    return workbook


# Обработчик команды для выгрузки данных в Excel
@dp.message_handler(commands=['export_orders'])
async def export_data(message: types.Message):
    if message.from_user.id not in [5837917794, 5958542955]:
        await message.reply('У вас нет доступа к этой команде.')
        return
    # Получение данных из базы данных
    cursor.execute("SELECT * FROM orders")
    orders = cursor.fetchall()
    # Создание файла Excel
    workbook = create_excel_file(orders)
    # Сохранение файла
    filename = 'orders.xlsx'
    workbook.save(filename)
    # Отправка файла пользователю
    with open(filename, 'rb') as file:
        await message.answer_document(file)
    # Удаление файла
    os.remove(filename)


def greeting_handler():
    """Регистрируем handlers для калькулятора"""
    dp.register_message_handler(greeting)  # Обработчик команды /start, он же пост приветствия
